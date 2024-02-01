import ftplib
import sqlite3
import os
import json
import os.path
from ftplib import FTP
import datetime
import time
from pathlib import Path
from sqlite3 import Error
import openpyxl
from pymodbus.constants import Endian
from pymodbus.payload import BinaryPayloadDecoder
from pymodbus.client.sync import ModbusSerialClient as ModbusClient

#install pymodbus2.2.0.rc2
#install openpyxl

duration = 0

class datalogger():
    def constructorFunc(self):
        self.real_datetime = datetime.datetime.now()
        self.real_time = self.real_datetime.strftime("%X")
        self.real_date = self.real_datetime.strftime("%x")
    
    def writeJsonFileError(self,text):
        try:
            data_error = {
                "error": text
            }
            json_object = json.dumps(data_error, indent = 1)
        
            with open("/home/songhong6/nuoc_ngam/jsonfile/error.json", "w") as outfile:
                outfile.write(json_object)
        except Exception as e:
            print("Loi: " + str(e))
    
    def readData(self,port,baud,type,address,id):
        try:
            client=ModbusClient(method='rtu', port=port, baudrate=baud,timeout=1,parity='N',strict=False,stopbits=1)
            client.connect()
            time.sleep(1)
            if type == "signed":
                response=client.read_holding_registers(int(address),2,unit=int(id))
                self.a = response.registers
                self.k = round(self.a[0]/100,2)
                client.close()
                return self.k
            if type == "long":
                response=client.read_holding_registers(int(address),2,unit=int(id))             
                value1 = response.registers[0]
                value2 = response.registers[1]               
                value = [value1,value2]   
                decoder = BinaryPayloadDecoder.fromRegisters(value, byteorder=Endian.Big)
                self.l = decoder.decode_32bit_int()
                client.close()
                return self.l
            if type == "float":      
                response=client.read_input_registers(int(address),2,unit=int(id))
                value1 = response.registers[0]
                value2 = response.registers[1]               
                value = [value1,value2]    
                decoder = BinaryPayloadDecoder.fromRegisters(value, byteorder=Endian.Big)
                self.p = decoder.decode_32bit_float()
                self.m = round(self.p,3)
                client.close()
                return self.m
            self.writeJsonFileError("None")   
        except Exception as e:
            print("Loi: " + str(e))
    
    def writeDataTxt(self,TT1,TT2,dLL,dMN):
        try:
            real_datetime = datetime.datetime.now()
            self.real_time = real_datetime.strftime("%X")
            self.real_date = real_datetime.strftime("%x")
            self.r_datetime =  '20' + self.real_date[6:8] + self.real_date[0:2] + self.real_date[3:5]  + self.real_time[0:2] + self.real_time[3:5] + "00" #self.real_time[6:8]
            content = 'LUULUONG' + '\t' + str(dLL) + '\t' + 'm3/h' + '\t'+ self.r_datetime + '\t' + '0'+ TT1 + '\n' + 'MUCNUOC' + '\t' + str(dMN) + '\t' + 'm' + '\t' + self.r_datetime + '\t' + '0' + TT2 + '\n' 
            file = open(self.filesave,'w')
            file.write(content)
            self.writeJsonFileError("None")      
        except Exception as e:
            print("Loi: " + str(e))
            self.writeJsonFileError("Error")
    
    def exportExcelFileFlow(self):
        try:
            book = openpyxl.Workbook()
            sheet = book.active
            self.cursor.execute('SELECT * FROM flow')
            d = self.cursor.fetchall()
            i = 0
            for row in d:
                sheet['A1'] = "Date"
                sheet['B1'] = "TOTAL_FLOW(m3)"
                i+=1
                j=1
                for col in row:
                    cell = sheet.cell(row = i+1, column = j)
                    cell.value = col
                    j += 1
            fileExcelFlow = "/home/songhong6/nuoc_ngam/excel/totalflow.xlsx"
            book.save(fileExcelFlow)
            self.writeJsonFileError("None")
            return fileExcelFlow
        except Exception as e:
            print("Loi: " + str(e))
            self.writeJsonFileError("Error")
    
    def exportExcelFile(self,year_start,month_start,day_start,year_fin,month_fin,day_fin):
        try:
            book = openpyxl.Workbook()
            sheet = book.active
            start = str(year_start+"/"+month_start+"/"+day_start)
            print("Dữ liệu được trích xuất trong ngày " + start)
            finish = str(year_fin+"/"+month_fin+"/"+day_fin)
            self.cursor.execute('SELECT * FROM database WHERE date BETWEEN (?) AND (?)',(start,finish))
            d = self.cursor.fetchall()
            i = 0
            for row in d:
                sheet['A1'] = "Date"
                sheet['B1'] = "Time"
                sheet['C1'] = "LUULUONG1(m3/h)"
                sheet['D1'] = "MUCNUOC1(m)"
                sheet['E1'] = "LUULUONG2(m3/h)"
                sheet['F1'] = "MUCNUOC2(m)"
                i+=1
                j=1
                for col in row:
                    cell = sheet.cell(row = i+1, column = j)
                    cell.value = col
                    j += 1
            fileExcel = "/home/songhong6/nuoc_ngam/excel/database.xlsx"
            book.save(fileExcel)
            self.writeJsonFileError("None")
            return fileExcel
        except Exception as e:
            print("Loi: " + str(e))
            self.writeJsonFileError("Error")
            
class ftpFile(datalogger):
    global filename, filesave, fileExcel
    def constructorFunc(self):
        self.ftpPort = 21 
    
    def connect(self, ftpHost, ftpUname, ftpPass):
        try:
            ftp = ftplib.FTP(ftpHost, ftpUname, ftpPass)
            ftp.encoding = "utf-8"
            self.ftp = FTP(ftpHost)
            self.ftp.login(ftpUname, ftpPass)
            self.writeJsonFileError("None")
        except Exception as e:
            print("Loi: " + str(e))
            self.writeJsonFileError("Error")
        
    def uploadFileFTP(self, filesave,name):
        try:
            self.ftp.cwd(name)
            year = "20"+ self.real_date[6:8]
            self.ftp.cwd(year)
            self.ftp.cwd(self.real_date[0:2])
            self.ftp.cwd(self.real_date[3:5])
            with open(filesave, "rb") as file:
                print(filesave)
                self.ftp.storbinary('STOR {}'.format(os.path.basename(filesave)),file,1024*1024)
            print('[Lưu file trên Server thành công]')
            self.ftp.quit()
            self.writeJsonFileError("None")
        except Exception as e:
            print("Loi: " + str(e))
            self.writeJsonFileError("Error")    
class dataStorage(ftpFile):
    proname = "TH_SHVINA_"
    sitename = "NUO_"
    def _init_(self):
        try:
            self.sqliteConnection = sqlite3.connect('SQLite_Database.db',
                                           detect_types=sqlite3.PARSE_DECLTYPES |
                                                        sqlite3.PARSE_COLNAMES)
            self.cursor = self.sqliteConnection.cursor()
        except Exception as e:
            print("Loi: " + str(e))
    
    def createTable(self):
        try:
            self.cursor.execute("""CREATE TABLE IF NOT EXISTS database (
                                       Date null,
                                       Time null,
                                       LUULUONG1 null,
                                       MUCNUOC1 null,
                                       LUULUONG2 null,
                                       MUCNUOC2 null
                                       )""")
        except Exception as e:
            print("Loi: " + str(e))
    
    def table(self):
        try:
            self.cursor.execute("""CREATE TABLE IF NOT EXISTS flow (
                                    Date null,
                                    TOTAL_FLOW null
                                    )""")
        except Exception as e:
            print("Loi: " + str(e))
    
    def addData(self,Date,Time,dLL1,dMN1,dLL2,dMN2):
        try:
        # them du lieu
            self.cursor.execute("""INSERT INTO database(Date,Time,LUULUONG1,MUCNUOC1,LUULUONG2,MUCNUOC2) VALUES (?,?,?,?,?,?)""",(Date,Time,dLL1,dMN1,dLL2,dMN2))
            self.sqliteConnection.commit()
        except Exception as e:
            print("Loi: " + str(e))

    def addDataFlow(self,Date,total):
        try:
            self.cursor.execute("""INSERT INTO flow(Date,TOTAL_FLOW) VALUES (?,?)""",(Date,total))
            self.sqliteConnection.commit()
        except Exception as e:
            print("Loi: " + str(e))
    
    def getData(self):  
        self.cursor.execute('SELECT * FROM database')
        print(self.cursor.fetchall())
    
    def getDataFlow(self):  
        self.cursor.execute('SELECT * FROM flow')
        print(self.cursor.fetchall())
    
    def createFolder(self,name):
        try:
            real_datetime = datetime.datetime.now()
            real_date = real_datetime.strftime("%x")
            if os.path.exists("/home/songhong6/nuoc_ngam/" + name) == False:
                os.mkdir("/home/songhong6/nuoc_ngam/" + name)
            if os.path.exists("/home/songhong6/nuoc_ngam/" + name + "/20" + real_date[6:8]) == False:  
                os.mkdir("/home/songhong6/nuoc_ngam/" + name + "/20" + real_date[6:8])
            if os.path.exists("/home/songhong6/nuoc_ngam/" + name + "/20" + real_date[6:8] + "/" + real_date[0:2]) == False: 
                os.mkdir("/home/songhong6/nuoc_ngam/" + name + "/20" + real_date[6:8] + "/" + real_date[0:2])
            if os.path.exists("/home/songhong6/nuoc_ngam/" + name + "/20" + real_date[6:8] + "/" + real_date[0:2] + "/" + real_date[3:5]) == False: 
                os.mkdir("/home/songhong6/nuoc_ngam/" + name + "/20" + real_date[6:8] + "/" + real_date[0:2] + "/" + real_date[3:5])
            
            bool0 = name in self.ftp.nlst()
            if bool0 == True:
                self.ftp.cwd(name)
                print("Thư mục " + name + " đã tồn tại trên Server")
            if bool0 == False:
                self.ftp.mkd(name)
                self.ftp.cwd(name)
                print("Thư mục " + name + " chưa tồn tại trên Server")
            folderName1 = "20" + real_date[6:8]
            folder1 = self.ftp.nlst()
            if folder1 != []:
                for c in folder1:
                    if c == folderName1:
                        bool1 = True
                        break
                    else:
                        bool1 = False
            else:
                bool1 = False
            if bool1 == True:
                self.ftp.cwd("20" + real_date[6:8])
                print("Thư mục 20" + real_date[6:8] + " đã tồn tại trên Server")
            if bool1 == False:
                self.ftp.mkd("20" + real_date[6:8])
                self.ftp.cwd("20" + real_date[6:8])
                print("Thư mục 20" + real_date[6:8] + " chưa tồn tại trên Server. Đã tạo thư mục 20" + real_date[6:8])
            folderName2 = real_date[0:2]
            folder2 = self.ftp.nlst()
            if folder2 != []:
                for c in folder2:
                    if c == folderName2:
                        bool2 = True
                        break
                    else:
                        bool2 = False
            else:
                bool2 = False
            if bool2 == True:
                self.ftp.cwd(real_date[0:2])
                print("Thư mục " + real_date[0:2] + " đã tồn tại trên Server")
            if bool2 == False:
                self.ftp.mkd(real_date[0:2])
                self.ftp.cwd(real_date[0:2])
                print("Thư mục " + real_date[0:2] + " chưa tồn tại trên Server. Đã tạo thư mục " + real_date[0:2])
            folderName3 = real_date[3:5]
            folder3 = self.ftp.nlst()
            if folder3 != []:
                for c in folder3:
                    if c == folderName3:
                        bool3 = True
                        break
                    else:
                        bool3 = False
            else:
                bool3 = False
            if bool3 == True:
                self.ftp.cwd(real_date[3:5])
                print("Thư mục " + str(real_date[3:5]) + " đã tồn tại trên Server")
            if bool3 == False:
                self.ftp.mkd(real_date[3:5])
                self.ftp.cwd(real_date[3:5])
                print("Thư mục " + str(real_date[3:5]) + " chưa tồn tại trên Server. Đã tạo thư mục " + str(real_date[3:5]))
            self.ftp.quit()
            self.writeJsonFileError("None")
        except Exception as e:
            print("Loi: " + str(e))
            self.writeJsonFileError("Error")
        
    def createFile(self,name,server,username,password,number):
        try:
            real_datetime = datetime.datetime.now()
            real_time = real_datetime.strftime("%X")
            real_date = real_datetime.strftime("%x")
            if os.path.exists("/home/songhong6/nuoc_ngam/" + name + "/20" + real_date[6:8] + "/" + real_date[0:2] + "/" + real_date[3:5]) == False:
                self.connect(server,username,password)
                self.createFolder(name)
                path = "/home/songhong6/nuoc_ngam/"+ name + "/20" + real_date[6:8] + "/" + real_date[0:2] + "/" + real_date[3:5]
            else:
                path = "/home/songhong6/nuoc_ngam/"+ name + "/20" + real_date[6:8] + "/" + real_date[0:2] + "/" + real_date[3:5]
            self.r_datetime =  '20' + real_date[6:8] + real_date[0:2] + real_date[3:5]  + real_time[0:2] + real_time[3:5] + "00"
            filename = os.path.join(path + "/" + self.proname + self.sitename + number + self.r_datetime  + '.txt')
            self.filesave = filename
            file = open(filename,"w")
            file.close()
            self.writeJsonFileError("None")
        except Exception as e:
            print("Loi: " + str(e))
            self.writeJsonFileError("Error")

class fileJson(dataStorage):
    server = ""
    username = ""
    password = ""
    
    def readJsonFileModbus(self):
        try:
            f = open('/home/songhong6/nuoc_ngam/jsonfile/dataModbus.json','r')
            d = f.read()
            obj=json.loads(d)
            port = str(obj['Port'])
            baud = int(obj['Baudrate'])
            f.close()
            return port, baud
        except Exception as e:
            print("Loi: " + str(e))
    
    def readValue(self):
        try:
            f = open('/home/songhong6/nuoc_ngam/jsonfile/sensor_values.json','r')
            d = f.read()
            obj=json.loads(d)
            luuluong1 = float(obj['LUULUONG1'])
            mucnuoc1 = float(obj['MUCNUOC1'])
            luuluong2 = float(obj['LUULUONG2'])
            mucnuoc2 = float(obj['MUCNUOC2'])
            f.close()
            return luuluong1, mucnuoc1, luuluong2, mucnuoc2
        except Exception as e:
            print("Loi: " + str(e))
    
    def readFlow(self):
        try:
            f = open('/home/songhong6/nuoc_ngam/jsonfile/Flow.json','r')
            d = f.read()
            obj=json.loads(d)
            total = int(obj['TOTAL'])
            pretotal = int(obj['PRETOTAL'])
            f.close()
            return total, pretotal
        except Exception as e:
            print("Loi: " + str(e))        
    
    def writeFlow(self,total,pretotal):
        try:
            data_flow = {
                "TOTAL" : total,
                "PRETOTAL" : pretotal   
            }
            json_object = json.dumps(data_flow, indent = 2)
        
            with open("/home/songhong6/nuoc_ngam/jsonfile/Flow.json", "w") as outfile:
                outfile.write(json_object)
        except Exception as e:
            print("Loi: " + str(e))
    
    def readJsonFileExcel(self):
        try:
            f = open('/home/songhong6/nuoc_ngam/jsonfile/exportExcel.json','r')
            d = f.read()
            obj=json.loads(d)
            export = str(obj['export'])
            f.close()
            return export
        except Exception as e:
            print("Loi: " + str(e))
    
    def readJsonFileExcelFlow(self):
        try:
            f = open('/home/songhong6/nuoc_ngam/jsonfile/exportExcel.json','r')
            d = f.read()
            obj=json.loads(d)
            exportFlow = str(obj['exportFlow'])
            f.close()
            return exportFlow
        except Exception as e:
            print("Loi: " + str(e))
    
    def readJsonSensorMode(self):
        try:
            f = open('/home/songhong6/nuoc_ngam/jsonfile/sensor_mode.json','r')
            d = f.read()
            obj=json.loads(d)
            self.LL1 = str(obj['Mode_LUULUONG1']['Mode'])
            self.MN1 = str(obj['Mode_MUCNUOC1']['Mode'])
            self.LL2 = str(obj['Mode_LUULUONG2']['Mode'])
            self.MN2 = str(obj['Mode_MUCNUOC2']['Mode'])
            f.close()
            return self.LL1,self.MN1,self.LL2,self.MN2
        except Exception as e:
            print("Loi: " + str(e))
    
    def writeJsonFileExcel(self,text):
        try:
            data_excel = {
                "export": text
            }
            json_object = json.dumps(data_excel, indent = 1)
        
            with open("/home/songhong6/nuoc_ngam/jsonfile/exportExcel.json", "w") as outfile:
                outfile.write(json_object)
        except Exception as e:
            print("Loi: " + str(e))
    
    def writeJsonFileExcelFlow(self,text):
        try:
            data_excel = {
                "exportFlow": text
            }
            json_object = json.dumps(data_excel, indent = 1)
        
            with open("/home/songhong6/nuoc_ngam/jsonfile/exportExcelFlow.json", "w") as outfile:
                outfile.write(json_object)
        except Exception as e:
            print("Loi: " + str(e))
    
    def readJsonFileFtp(self):
        try:
            f = open('/home/songhong6/nuoc_ngam/jsonfile/dataFtp.json','r')
            d = f.read()
            obj=json.loads(d)
            self.server = str(obj['Server'])
            self.username = str(obj['Username'])
            self.password = str(obj['Password'])
            f.close()
            return self.server, self.username, self.password
        except Exception as e:
            print("Loi: " + str(e))
    
    def writeJsonFile(self,LL1,MN1,LL2,MN2):
        try:
            data = {
                "LUULUONG1" : str(LL1),
                "MUCNUOC1" : str(MN1),
                "LUULUONG2" : str(LL2),
                "MUCNUOC2" : str(MN2)   
            }
            json_object = json.dumps(data, indent = 4)
        
            with open("/home/songhong6/nuoc_ngam/jsonfile/sensor_values.json", "w") as outfile:
                outfile.write(json_object)
        except Exception as e:
            print("Loi: " + str(e))
            
class configuration(fileJson):
    a = [0]*3
    modbus_ftp = fileJson()
    ftp = ftplib.FTP()
    ftp.encoding = "utf-8"
    sqliteConnection = sqlite3.connect('SQLite_Database.db',
                                           detect_types=sqlite3.PARSE_DECLTYPES |
                                                        sqlite3.PARSE_COLNAMES)
    cursor = sqliteConnection.cursor()
    def configModbus(self):
        try:
            a = self.modbus_ftp.readJsonFileModbus()
            self.port = a[0]
            self.baud = a[1]
            self.writeJsonFileError("None")
        except Exception as e:
            print("Loi: " + str(e))
            self.writeJsonFileError("Error")
        
    def configFTP(self):
        try:
            a = self.modbus_ftp.readJsonFileFtp()
            self.server = a[0]
            self.username = a[1]
            self.password = a[2]
            self.writeJsonFileError("None")
        except Exception as e:
            print("Loi: " + str(e))
            self.writeJsonFileError("Error")

obj1 = configuration()
database = dataStorage()
name1 = "Gieng1"
name2 = "Gieng2"
while True:
    check = 0
    arr = obj1.readValue()
    list_arr = list(arr)
    arr1 = obj1.readFlow()
    list_arr1 = list(arr1)
    real_datetime = datetime.datetime.now()
    real_time = real_datetime.strftime("%X")
    real_date = real_datetime.strftime("%x")
    time1 = "20" + real_date[6:8] + "/" + real_date[0:2] + "/" + real_date[3:5]   
    obj1.configFTP()
    obj1.configModbus()
    database._init_()
    obj1.createTable()
    obj1.table()
    obj1.connect(obj1.server,obj1.username,obj1.password)
    obj1.createFolder(name1)
    obj1.connect(obj1.server,obj1.username,obj1.password)
    obj1.createFolder(name2)
    while check == 0:
        dLL1 = obj1.readData(obj1.port,obj1.baud,"signed","11","1")
        dMN1 = obj1.readData(obj1.port,obj1.baud,"float","8","3")
        dLL2 = obj1.readData(obj1.port,obj1.baud,"signed","11","2")
        dMN2 = obj1.readData(obj1.port,obj1.baud,"float","8","4")
        if dLL1 is None:
            dLL1 = list_arr[0]
        elif dLL1<0 or dLL1 >100:
            dLL1 = list_arr[0]
        else:
            list_arr[0] = dLL1
        if dMN1 is None:
            dMN1 = list_arr[1]
        elif dMN1<0 or dMN1 >100:
            dMN1 = list_arr[1]
        else:
            list_arr[1] = dMN1    
        if dLL2 is None:
            dLL2 = list_arr[2]
        elif dLL2<0 or dLL2 >100:
            dLL2 = list_arr[2]
        else:
            list_arr[2] = dLL2
        if dMN2 is None:
            dMN2 = list_arr[3]
        elif dMN2<0 or dMN2 >100:
            dMN2 = list_arr[3]
        else:
            list_arr[3] = dMN2
        obj1.writeJsonFile(dLL1,dMN1,dLL2,dMN2)
        real_datetime = datetime.datetime.now()
        real_time = real_datetime.strftime("%X")
        real_date = real_datetime.strftime("%x")
        r_time =  '20' + real_date[6:8] + real_date[0:2] + real_date[3:5]  + real_time[0:2] + real_time[3:5] + "00"   
        filepath = "/home/songhong6/nuoc_ngam/"+ name2 + "/20" + real_date[6:8] + "/" + real_date[0:2] + "/" + real_date[3:5] + "/" + obj1.proname + obj1.sitename + "002_" + r_time  + '.txt'
        if int(real_time[3:5])%5==0:
            if os.path.isfile(filepath):
                check = 0
            else:
                check = 1
        r1_time =  '20' + real_date[6:8] + real_date[0:2] + real_date[3:5]  + "070000"
        filepath1 = "/home/songhong6/nuoc_ngam/"+ name2 + "/20" + real_date[6:8] + "/" + real_date[0:2] + "/" + real_date[3:5] + "/" + obj1.proname + obj1.sitename + "002_" + r1_time  + '.txt'
        if real_time[0:2] == '07':
            if os.path.isfile(filepath1) == False:
                pre_Total = obj1.readData(obj1.port,obj1.baud,"long","2","2")  #address-id
                answer = int(pre_Total) - int(list_arr1[1])
                obj1.writeFlow(answer,pre_Total)
                obj1.addDataFlow(time1,answer)
        export = obj1.readJsonFileExcel()
        if export == 'on':
            obj1.exportExcelFile(year_start=str(2023),month_start=str(real_date[0:2]),day_start=str(real_date[3:5]),year_fin=str(2023),month_fin = str(real_date[0:2]),day_fin=str(real_date[3:5]))  
            obj1.writeJsonFileExcel('off')
        exportFlow = obj1.readJsonFileExcelFlow()
        if exportFlow == 'on':
            obj1.exportExcelFileFlow()  
            obj1.writeJsonFileExcelFlow('off')
        time.sleep(2)     

    real_datetime = datetime.datetime.now()
    real_time = real_datetime.strftime("%X")
    real_date = real_datetime.strftime("%x")
    time1 = "20" + real_date[6:8] + "/" + real_date[0:2] + "/" + real_date[3:5]    
    obj1.createFile(name1,obj1.server,obj1.username,obj1.password,"001_")
    obj1.readJsonSensorMode() 
    obj1.writeDataTxt(obj1.LL1,obj1.MN1,dLL=dLL1,dMN=dMN1)
    obj1.connect(obj1.server,obj1.username,obj1.password)
    obj1.uploadFileFTP(obj1.filesave,name1)
    obj1.createFile(name2,obj1.server,obj1.username,obj1.password,"002_")
    obj1.readJsonSensorMode()
    obj1.writeDataTxt(obj1.LL2,obj1.MN2,dLL=dLL2,dMN=dMN2)
    obj1.connect(obj1.server,obj1.username,obj1.password)   
    obj1.uploadFileFTP(obj1.filesave,name2)
    obj1.addData(Date=time1,Time=real_time,dLL1=dLL1,dMN1=dMN1,dLL2=dLL2,dMN2=dMN2)
