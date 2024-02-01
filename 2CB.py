import ftplib
import sqlite3
import os
import json
import os.path
from ftplib import FTP
import datetime
import time
from pathlib import Path
from sqlite3 import Error
import openpyxl
from pymodbus.constants import Endian
from pymodbus.payload import BinaryPayloadDecoder
from pymodbus.client.sync import ModbusSerialClient as ModbusClient
import random

#install pymodbus
#install openpyxl

duration = 0

class datalogger():
    def constructorFunc(self):
        self.real_datetime = datetime.datetime.now()
        self.real_time = self.real_datetime.strftime("%X")
        self.real_date = self.real_datetime.strftime("%x")
    
    #def dieukhientulaymau():
    def writeJsonFileError(self,text):
        data_error = {
            "error": text
        }
        json_object = json.dumps(data_error, indent = 1)
        
        with open("/home/songhong6/nuoc_ngam/jsonfile/error.json", "w") as outfile:
            outfile.write(json_object)
    
    def readData(self,port,baud,typed,address,id):
        try:
            client=ModbusClient(method='rtu', port=port, baudrate=baud,timeout=1,parity='N',strict=False,stopbits=1)
            client.connect()
            time.sleep(1)
            if typed == "signed":
                response=client.read_holding_registers(int(address),2,unit=int(id))
                self.a = response.registers
                self.k = round(self.a[0]/100,2)
                client.close()
                return self.k
            if typed == "long":
                response=client.read_holding_registers(int(address),2,unit=int(id))             
                value1 = response.registers[0]
                value2 = response.registers[1]               
                value = [value1,value2]   
                decoder = BinaryPayloadDecoder.fromRegisters(value, byteorder=Endian.Big)
                self.l = decoder.decode_32bit_int()
                client.close()
                return self.l
            if typed == "float":      
                response=client.read_input_registers(int(address),2,unit=int(id))
                value1 = response.registers[0]
                value2 = response.registers[1]               
                value = [value1,value2]    
                decoder = BinaryPayloadDecoder.fromRegisters(value, byteorder=Endian.Big)
                self.p = decoder.decode_32bit_float()
                self.m = round(self.p,3)
                client.close()
                return self.m
            self.writeJsonFileError("None")
            
        except:
            self.writeJsonFileError("Error")
    
    def writeDataTxt(self,dLL,dMN):
        try:
            real_datetime = datetime.datetime.now()
            self.real_time = real_datetime.strftime("%X")
            self.real_date = real_datetime.strftime("%x")
            self.r_datetime =  '20' + self.real_date[6:8] + self.real_date[0:2] + self.real_date[3:5]  + self.real_time[0:2] + self.real_time[3:5] + "00" #self.real_time[6:8]
            content = 'LUULUONG' + '\t' + str(dLL) + '\t' + 'm3/h' +'\t'+ self.r_datetime + '\t' + '00'+ '\n' + 'MUCNUOC' + '\t' + str(dMN) + '\t' + 'm' +'\t' + self.r_datetime + '\t' + '00' + '\n' 
            file = open(self.filesave,'w')
            file.write(content)
            self.writeJsonFileError("None")
            
        except:
            print("Lỗi ghi dữ liệu vào file txt")
            self.writeJsonFileError("Error")
    
    def exportExcelFileFlow(self):
        try:
            book = openpyxl.Workbook()
            sheet = book.active
            self.cursor.execute('SELECT * FROM flow')
            d = self.cursor.fetchall()
            i = 0
            for row in d:
                sheet['A1'] = "Date"
                sheet['B1'] = "Time"
                sheet['C1'] = "TONGGIENG(m3)"
                i+=1
                j=1
                for col in row:
                    cell = sheet.cell(row = i+1, column = j)
                    cell.value = col
                    j += 1
            fileExcelFlow = "/home/songhong6/nuoc_ngam/excel/Totalflow.xlsx"
            book.save(fileExcelFlow)
            self.writeJsonFileError("None")
            return fileExcelFlow
        except Exception as e:
            print("Loi: " + str(e))
            self.writeJsonFileError("Error")
    
    def exportExcelFile(self,year_start,month_start,day_start,year_fin,month_fin,day_fin):
        try:
            book = openpyxl.Workbook()
            sheet = book.active
            start = str(year_start+"/"+month_start+"/"+day_start)
            print("Dữ liệu được trích xuất trong ngày " + start)
            finish = str(year_fin+"/"+month_fin+"/"+day_fin)
            self.cursor.execute('SELECT * FROM database WHERE date BETWEEN (?) AND (?)',(start,finish))
            d = self.cursor.fetchall()
            i = 0
            for row in d:
                sheet['A1'] = "Date"
                sheet['B1'] = "Time"
                sheet['C1'] = "LUULUONG(m3/h)"
                sheet['D1'] = "MUCNUOC(m)"
                i+=1
                j=1
                for col in row:
                    cell = sheet.cell(row = i+1, column = j)
                    cell.value = col
                    j += 1
            fileExcel = "/home/songhong6/nuoc_ngam/excel/database.xlsx"
            book.save(fileExcel)
            self.writeJsonFileError("None")
            return fileExcel
        except:
            print("Lỗi khi xuất file Excel")
            self.writeJsonFileError("Error")
            
class ftpFile(datalogger):
    global filename, filesave, fileExcel
    def constructorFunc(self):
        self.ftpPort = 21 
    
    def connect(self, ftpHost, ftpUname, ftpPass):
        try:
            ftp = ftplib.FTP(ftpHost, ftpUname, ftpPass)
            ftp.encoding = "utf-8"
            self.ftp = FTP(ftpHost)
            self.ftp.login(ftpUname, ftpPass)
            self.writeJsonFileError("None")
        except:
            print("Lỗi khi kết nối tới Server. Kiểm tra mạng internet")
            self.writeJsonFileError("Error")
        
    def uploadFileFTP(self, filesave,name):
        try:
            self.ftp.cwd(name)
            year = "20"+ self.real_date[6:8]
            self.ftp.cwd(year)
            self.ftp.cwd(self.real_date[0:2])
            self.ftp.cwd(self.real_date[3:5])
            with open(filesave, "rb") as file:
                print(filesave)
                self.ftp.storbinary('STOR {}'.format(os.path.basename(filesave)),file,1024*1024)
            print('[Lưu file trên Server thành công]')
            self.ftp.quit()
            self.writeJsonFileError("None")
        except:
            print("Không thể gửi file tới Server")
            self.writeJsonFileError("Error")    
class dataStorage(ftpFile):
    proname = "ND_QUYNHAT_"
    sitename = "GIENG10_"
    def _init_(self):
        try:
            self.sqliteConnection = sqlite3.connect('SQLite_Database.db',
                                           detect_types=sqlite3.PARSE_DECLTYPES |
                                                        sqlite3.PARSE_COLNAMES)
            self.cursor = self.sqliteConnection.cursor()
        except:
            print("Không thể kết nối tới SQLite database")
    
    def createTable(self):
        try:
            self.cursor.execute("""CREATE TABLE IF NOT EXISTS database (
                                       Date null,
                                       Time null,
                                       LUULUONG null,
                                       MUCNUOC null
                                       )""")
        except:
            print("Lỗi tạo bảng database")
    
    def table(self):
        try:
            self.cursor.execute("""CREATE TABLE IF NOT EXISTS flow (
                                    Date null,
                                    Time null,
                                    TONGGIENG null
                                    )""")
        except Exception as e:
            print("Loi: " + str(e))
    
    def addData(self,Date,Time,dLL,dMN):
        try:
        # them du lieu
            self.cursor.execute("""INSERT INTO database(Date,Time,LUULUONG,MUCNUOC) VALUES (?,?,?,?)""",(Date,Time,dLL,dMN))
            self.sqliteConnection.commit()
        except:
            print("Lỗi thêm dữ liệu vao Database")
    
    def addDataFlow(self,Date,Time,dTG):
        try:
            self.cursor.execute("""INSERT INTO flow(Date,Time,TONGGIENG) VALUES (?,?,?)""",(Date,Time,dTG))
            self.sqliteConnection.commit()
        except Exception as e:
            print("Loi: " + str(e))
    
    def getData(self):  
        self.cursor.execute('SELECT * FROM database')
        print(self.cursor.fetchall())
    
    def createFolder(self,name):
        try:
            real_datetime = datetime.datetime.now()
            real_date = real_datetime.strftime("%x")
            if os.path.exists("/home/songhong6/nuoc_ngam/" + name) == False:
                os.mkdir("/home/songhong6/nuoc_ngam/" + name)
            if os.path.exists("/home/songhong6/nuoc_ngam/" + name + "/20" + real_date[6:8]) == False:  
                os.mkdir("/home/songhong6/nuoc_ngam/" + name + "/20" + real_date[6:8])
            if os.path.exists("/home/songhong6/nuoc_ngam/" + name + "/20" + real_date[6:8] + "/" + real_date[0:2]) == False: 
                os.mkdir("/home/songhong6/nuoc_ngam/" + name + "/20" + real_date[6:8] + "/" + real_date[0:2])
            if os.path.exists("/home/songhong6/nuoc_ngam/" + name + "/20" + real_date[6:8] + "/" + real_date[0:2] + "/" + real_date[3:5]) == False: 
                os.mkdir("/home/songhong6/nuoc_ngam/" + name + "/20" + real_date[6:8] + "/" + real_date[0:2] + "/" + real_date[3:5])
            
            bool0 = name in self.ftp.nlst()
            if bool0 == True:
                self.ftp.cwd(name)
                print("Thư mục " + name + " đã tồn tại trên Server")
            if bool0 == False:
                self.ftp.mkd(name)
                self.ftp.cwd(name)
                print("Thư mục " + name + " chưa tồn tại trên Server")
            folderName1 = "20" + real_date[6:8]
            folder1 = self.ftp.nlst()
            if folder1 != []:
                for c in folder1:
                    if c == folderName1:
                        bool1 = True
                        break
                    else:
                        bool1 = False
            else:
                bool1 = False
            if bool1 == True:
                self.ftp.cwd("20" + real_date[6:8])
                print("Thư mục 20" + real_date[6:8] + " đã tồn tại trên Server")
            if bool1 == False:
                self.ftp.mkd("20" + real_date[6:8])
                self.ftp.cwd("20" + real_date[6:8])
                print("Thư mục 20" + real_date[6:8] + " chưa tồn tại trên Server. Đã tạo thư mục 20" + real_date[6:8])
            folderName2 = real_date[0:2]
            folder2 = self.ftp.nlst()
            if folder2 != []:
                for c in folder2:
                    if c == folderName2:
                        bool2 = True
                        break
                    else:
                        bool2 = False
            else:
                bool2 = False
            if bool2 == True:
                self.ftp.cwd(real_date[0:2])
                print("Thư mục " + real_date[0:2] + " đã tồn tại trên Server")
            if bool2 == False:
                self.ftp.mkd(real_date[0:2])
                self.ftp.cwd(real_date[0:2])
                print("Thư mục " + real_date[0:2] + " chưa tồn tại trên Server. Đã tạo thư mục " + real_date[0:2])
            folderName3 = real_date[3:5]
            folder3 = self.ftp.nlst()
            if folder3 != []:
                for c in folder3:
                    if c == folderName3:
                        bool3 = True
                        break
                    else:
                        bool3 = False
            else:
                bool3 = False
            if bool3 == True:
                self.ftp.cwd(real_date[3:5])
                print("Thư mục " + str(real_date[3:5]) + " đã tồn tại trên Server")
            if bool3 == False:
                self.ftp.mkd(real_date[3:5])
                self.ftp.cwd(real_date[3:5])
                print("Thư mục " + str(real_date[3:5]) + " chưa tồn tại trên Server. Đã tạo thư mục " + str(real_date[3:5]))
            self.ftp.quit()
            self.writeJsonFileError("None")
        except:
            print("Lỗi khi tạo Folder")
            self.writeJsonFileError("Error")
        
    def createFile(self,name,server,username,password):
        try:
            real_datetime = datetime.datetime.now()
            real_time = real_datetime.strftime("%X")
            real_date = real_datetime.strftime("%x")
            if os.path.exists("/home/songhong6/nuoc_ngam/" + name + "/20" + real_date[6:8] + "/" + real_date[0:2] + "/" + real_date[3:5]) == False:
                self.connect(server,username,password)
                self.createFolder(name)
                path = "/home/songhong6/nuoc_ngam/"+ name + "/20" + real_date[6:8] + "/" + real_date[0:2] + "/" + real_date[3:5]
            else:
                path = "/home/songhong6/nuoc_ngam/"+ name + "/20" + real_date[6:8] + "/" + real_date[0:2] + "/" + real_date[3:5]
            self.r_datetime =  '20' + real_date[6:8] + real_date[0:2] + real_date[3:5]  + real_time[0:2] + real_time[3:5] + "00"
            filename = os.path.join(path + "/" + self.proname + self.sitename + self.r_datetime  + '.txt')
            self.filesave = filename
            file = open(filename,"w")
            file.close()
            self.writeJsonFileError("None")
        except:
            print("Lỗi khi tạo file text")
            self.writeJsonFileError("Error")

class fileJson(dataStorage):
    server = ""
    username = ""
    password = ""
    
    def readJsonFileModbus(self):
        try:
            f = open('/home/songhong6/nuoc_ngam/jsonfile/dataModbus.json','r')
            d = f.read()
            obj=json.loads(d)
            port = str(obj['Port'])
            baud = int(obj['Baudrate'])
            f.close()
            return port, baud
        except:
            print("Không thể đọc dữ liệu cấu hình Modbus RTU")
    
    def readValue(self):
        try:
            f = open('/home/songhong6/nuoc_ngam/jsonfile/sensor_values.json','r')
            d = f.read()
            obj=json.loads(d)
            luuluong = float(obj['LUULUONG'])
            mucnuoc = float(obj['MUCNUOC'])
            f.close()
            return luuluong, mucnuoc
        except Exception as e:
            print("Loi: " + str(e))
    
    def readFlow(self):
        try:
            f = open('/home/songhong6/nuoc_ngam/jsonfile/Flow.json','r')
            d = f.read()
            obj=json.loads(d)
            tg = float(obj['TongGieng'])
            f.close()
            return tg
        except Exception as e:
            print("Loi: " + str(e))  
    
    def writeFlow(self,tg):
        try:
            data_flow = {
                "TongGieng" : str(tg)
            }
            json_object = json.dumps(data_flow, indent = 1)
        
            with open("/home/songhong6/nuoc_ngam/jsonfile/Flow.json", "w") as outfile:
                outfile.write(json_object)
        except Exception as e:
            print("Loi: " + str(e))
                
    def readJsonFileExcel(self):
        try:
            f = open('/home/songhong6/nuoc_ngam/jsonfile/exportExcel.json','r')
            d = f.read()
            obj=json.loads(d)
            export = str(obj['export'])
            f.close()
            return export
        except:
            print("Không nhận được yêu cầu xuất file Excel")
    
    def readJsonFileExcelFlow(self):
        try:
            f = open('/home/songhong6/nuoc_ngam/jsonfile/exportExcelFlow.json','r')
            d = f.read()
            obj=json.loads(d)
            exportFlow = str(obj['exportFlow'])
            f.close()
            return exportFlow
        except Exception as e:
            print("Loi: " + str(e))
    
    def readJsonSensorMode(self):
        try:
            f = open('/home/songhong6/nuoc_ngam/jsonfile/sensor_mode.json','r')
            d = f.read()
            obj=json.loads(d)
            self.LL = str(obj['Mode_LUULUONG']['Mode'])
            self.MN = str(obj['Mode_MUCNUOC']['Mode'])
            f.close()
            return self.LL,self.MN
        except:
            print("Không nhận được trạng thái của cảm biến")
    
    def writeJsonFileExcel(self):
        try:
            data_excel = {
                "export": "off"
            }
            json_object = json.dumps(data_excel, indent = 1)
        
            with open("/home/songhong6/nuoc_ngam/jsonfile/exportExcel.json", "w") as outfile:
                outfile.write(json_object)
        except:
            print("Lỗi trạng thái xuất file Excel")
    
    def writeJsonFileExcelFlow(self):
        try:
            data_excel_flow = {
                "exportFlow": "off"
            }
            json_object_flow = json.dumps(data_excel_flow, indent = 1)
        
            with open("/home/songhong6/nuoc_ngam/jsonfile/exportExcelFlow.json", "w") as outfile:
                outfile.write(json_object_flow)
        except Exception as e:
            print("Loi: " + str(e))
    
    def readJsonFileFtp(self):
        try:
            f = open('/home/songhong6/nuoc_ngam/jsonfile/dataFtp.json','r')
            d = f.read()
            obj=json.loads(d)
            self.server = str(obj['Server'])
            self.username = str(obj['Username'])
            self.password = str(obj['Password'])
            f.close()
            return self.server, self.username, self.password
        except:
            print("Không nhận được dữ liệu cấu hình FTP")
    
    def writeJsonFile(self,LL,MN):
        try:
            data = {
                "LUULUONG" : str(LL),
                "MUCNUOC" : str(MN)  
            }
            json_object = json.dumps(data, indent = 4)
        
            with open("/home/songhong6/nuoc_ngam/jsonfile/sensor_values.json", "w") as outfile:
                outfile.write(json_object)
        except:
            print("Không cập nhật được giá trị tới màn hình")

class configuration(fileJson):
    a = [0]*3
    modbus_ftp = fileJson()
    ftp = ftplib.FTP()
    ftp.encoding = "utf-8"
    sqliteConnection = sqlite3.connect('SQLite_Database.db',
                                           detect_types=sqlite3.PARSE_DECLTYPES |
                                                        sqlite3.PARSE_COLNAMES)
    cursor = sqliteConnection.cursor()
    def configModbus(self):
        try:
            a = self.modbus_ftp.readJsonFileModbus()
            self.port = a[0]
            self.baud = a[1]
            self.writeJsonFileError("None")
        except:
            print("Chưa cấu hình Modbus RTU")
            self.writeJsonFileError("Error")
        
    def configFTP(self):
        try:
            a = self.modbus_ftp.readJsonFileFtp()
            self.server = a[0]
            self.username = a[1]
            self.password = a[2]
            self.writeJsonFileError("None")
        except:
            print("Chưa cấu hình FTP server")
            self.writeJsonFileError("Error")

obj1 = configuration()
obj2 = configuration()
database = dataStorage()
name1 = "Quy Nhat Gieng 10"

arr = obj1.readValue()
list_arr = list(arr)
real_datetime = datetime.datetime.now()
real_time = real_datetime.strftime("%X")
real_date = real_datetime.strftime("%x")
time1 = "20" + real_date[6:8] + "/" + real_date[0:2] + "/" + real_date[3:5]     
obj1.configFTP()
obj1.configModbus()
database._init_()
obj1.createTable()
obj2.table()
obj1.connect(obj1.server,obj1.username,obj1.password)
obj1.createFolder(name1)
t1 = time.time()
while(duration<298):
    dLL = abs(float(obj1.readData(obj1.port,obj1.baud,"float","4112","1")))
    dMN = abs(float(obj1.readData(obj1.port,obj1.baud,"float","4112","3")))
    pre_Total = obj2.readData(obj1.port,obj1.baud,"long","2","1")  #address-id   
    print(dLL)
    print(dMN)
    print("*****")
    if dLL is None:
        dLL = list_arr[0]
    elif dLL<0 or dLL >100:
        dLL = list_arr[0]
    else:
        list_arr[0] = dLL
    if dMN is None:
        dMN = list_arr[1]
    elif dMN<0 or dMN >100:
        dMN = list_arr[1]
    else:
        list_arr[1] = dMN    
    
    print("*****")
    obj1.writeJsonFile(dLL,dMN)
    real_datetime = datetime.datetime.now()
    real_time = real_datetime.strftime("%X")
    real_date = real_datetime.strftime("%x")
    export = obj1.readJsonFileExcel()
    if export == 'on':
        obj1.exportExcelFile(year_start=str(2023),month_start=str(real_date[0:2]),day_start=str(real_date[3:5]),year_fin=str(2023),month_fin = str(real_date[0:2]),day_fin=str(real_date[3:5]))  
        obj1.writeJsonFileExcel()
    #exportFlow = obj1.readJsonFileExcelFlow()
    #if exportFlow == 'on':
    #   obj1.exportExcelFileFlow()  
    #   obj1.writeJsonFileExcelFlow()
    duration = time.time() - t1
    time.sleep(2)     

real_datetime = datetime.datetime.now()
real_time = real_datetime.strftime("%X")
real_date = real_datetime.strftime("%x")
time1 = "20" + real_date[6:8] + "/" + real_date[0:2] + "/" + real_date[3:5]    
obj1.createFile(name1,obj1.server,obj1.username,obj1.password)
obj1.writeDataTxt(dLL=dLL,dMN=dMN)
obj1.connect(obj1.server,obj1.username,obj1.password)
obj1.addData(Date=time1,Time=real_time,dLL=dLL,dMN=dMN)
duration = 0
